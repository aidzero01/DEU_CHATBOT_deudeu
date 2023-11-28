from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup  
#찾을 객체의 위치정보 xpath
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from openpyxl import load_workbook
import string
#사용전에 pip install selenium,pip install webdriver-manager, pip install bs4 터미널로 설치
def dynamic():

    load_xlsx = load_workbook('C:/teamproject/front_end/excel/file.xlsx', 
                       read_only=False,
                       keep_vba=False,
                       data_only=True,
                       keep_links=True)
    load_sheet=load_xlsx['menu']
    
    #한식 일주일 메뉴를 담아줄 배열
    Korean_morning=[]
    Korean_lunch=[]
    Korean_dinner=[]

    #크롬 드라이브 설치
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    #내가 크롤링 해 올 주소
    driver.get("https://dorm.deu.ac.kr/hyomin/60/6050.kmc")

    #주소에 특정요일의 메뉴가 아닌 전체메뉴를 선택하기 위해 만든 코드
    driver.find_element(By.XPATH,'//*[@id="tabDayA"]').click()

    #웹페이지 로딩시간 기다려주는 코드. 3초까지 기다려줌
    #이 방식은 암시적 방법임.
    driver.implicitly_wait(3)
    
    #반복문을 통해 배열에 각 요일의 값의 메뉴를 넣음
    for i in range(1,8):

        #각 요일 한식메뉴 id값을 찾음
        Kmorning = '//*[@id="fo_menu_mor{}"]'.format(i)
        Klunch = '//*[@id="fo_menu_lun{}"]'.format(i)
        Kdinner = '//*[@id="fo_menu_eve{}"]'.format(i)

        #id값을 이용해 값을 찾는 과정
        km = driver.find_element(By.XPATH,Kmorning)
        kl = driver.find_element(By.XPATH,Klunch)
        kd = driver.find_element(By.XPATH,Kdinner)
        
        #배열에 각 요일별 한식 아침메뉴 넣음
        Korean_morning.append("-"+km.text)
        Korean_lunch.append("-"+kl.text)
        Korean_dinner.append("-"+kd.text)

    
    for i in range(0,7):
        Korean_morning[i] = Korean_morning[i].replace(" ","\n-")
        Korean_lunch[i] = Korean_lunch[i].replace(" ","\n-")
        Korean_dinner[i] = Korean_dinner[i].replace(" ","\n-")


    for i in range(1,8):
        load_sheet.cell(row=10, column=i+1, value=Korean_morning[i-1])
        load_sheet.cell(row=11, column=i+1, value=Korean_lunch[i-1])
        load_sheet.cell(row=12, column=i+1, value=Korean_dinner[i-1])
        
        
    load_xlsx.save("C:/teamproject/front_end/excel/file.xlsx")
    

dynamic()