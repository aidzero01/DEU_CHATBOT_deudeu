from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup  
#찾을 객체의 위치정보 xpath
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from openpyxl import load_workbook
#사용전에 pip install selenium,pip install webdriver-manager, pip install bs4 터미널로 설치
def dynamic():
    #엑셀 저장하기 위한 코드코드
    load_xlsx = load_workbook('C:/teamproject/front_end/excel/file.xlsx', 
                       read_only=False,
                       keep_vba=False,
                       data_only=True,
                       keep_links=True)
    load_sheet = load_xlsx.active
    
    #크롬 드라이브 설치
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    #내가 크롤링 해 올 주소
    driver.get("https://www.google.co.kr/search?q=%EB%82%A0%EC%94%A8&sca_esv=584981298&sxsrf=AM9HkKkYzuohRk7UpzH2syIqnhqT3ODMyw%3A1700802211003&source=hp&ei=oi5gZbC2O5DnwQPoqpWIAg&iflsig=AO6bgOgAAAAAZWA8s7PKx0D6G03lfz-hpZM8L7-2pdcd&ved=0ahUKEwiwj_n87duCAxWQc3AKHWhVBSEQ4dUDCAo&uact=5&oq=%EB%82%A0%EC%94%A8&gs_lp=Egdnd3Mtd2l6IgbrgqDslKgyDRAjGIAEGIoFGCcYnQIyCxAAGIAEGLEDGIMBMgsQABiABBixAxiDATILEAAYgAQYsQMYgwEyBRAAGIAEMgUQABiABDILEAAYgAQYsQMYgwEyCxAuGIAEGLEDGIMBMgoQABiABBiKBRhDMgUQABiABEjsBlBqWLQFcAF4AJABAJgBbaABlwOqAQMxLjO4AQPIAQD4AQGoAgrCAgcQIxjqAhgnwgIKECMYgAQYigUYJ8ICFhAuGIAEGBQYhwIYsQMYgwEYxwEY0QPCAhAQABiABBgUGIcCGLEDGIMBwgIREC4YgAQYsQMYgwEYxwEY0QPCAgQQABgDwgIEECMYJw&sclient=gws-wiz")

    #웹페이지 로딩시간 기다려주는 코드. 3초까지 기다려줌
    #이 방식은 암시적 방법임.
    driver.implicitly_wait(3)
    

    #각 요일 한식메뉴 id값을 찾음
    Wt = '//*[@id="wob_tm"]'
    #id값을 이용해 값을 찾는 과정
    weather = driver.find_element(By.XPATH,Wt)  
    
    #날씨 기온정보가 저장될 엑셀 위치 정하면 됨  
    load_sheet.cell(row=7, column=1, value=weather.text)
    load_sheet.cell(row=7, column=1, value=weather.text)
            
            
    load_xlsx.save("C:/teamproject/front_end/excel/file.xlsx")
        

dynamic()


